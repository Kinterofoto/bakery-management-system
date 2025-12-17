"""World Office Excel Generator Service.

Generates Excel files compatible with World Office import format.
Replicates the exact logic from use-world-office-export.ts hook.
"""

import logging
from io import BytesIO
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional

logger = logging.getLogger(__name__)

# Column headers in exact order (44 columns total)
EXCEL_COLUMNS = [
    "Encab: Empresa",
    "Encab: Tipo Documento",
    "Encab: Prefijo",
    "Encab: Documento Número",
    "Encab: Fecha",
    "Encab: Tercero Interno",
    "Encab: Tercero Externo",
    "Encab: Nota",
    "Encab: FormaPago",
    "Encab: Fecha Entrega",
    "Encab: Prefijo Documento Externo",
    "Encab: Número_Documento_Externo",
    "Encab: Verificado",
    "Encab: Anulado",
    "Encab: Personalizado 1",
    "Encab: Personalizado 2",
    "Encab: Personalizado 3",
    "Encab: Personalizado 4",
    "Encab: Personalizado 5",
    "Encab: Personalizado 6",
    "Encab: Personalizado 7",
    "Encab: Personalizado 8",
    "Encab: Personalizado 9",
    "Encab: Personalizado 10",
    "Encab: Personalizado 11",
    "Encab: Personalizado 12",
    "Encab: Personalizado 13",
    "Encab: Personalizado 14",
    "Encab: Personalizado 15",
    "Encab: Sucursal",
    "Encab: Clasificación",
    "Detalle: Producto",
    "Detalle: Bodega",
    "Detalle: UnidadDeMedida",
    "Cantidad",
    "Detalle: IVA",
    "Detalle: Valor Unitario",
    "Detalle: Descuento",
    "Detalle: Vencimiento",
    "Detalle: Nota",
    "Detalle: Centro costos",
    "Detalle: Personalizado1",
    "Detalle: Personalizado2",
    "Detalle: Personalizado3",
    "Detalle: Personalizado4",
    "Detalle: Personalizado5",
    "Detalle: Personalizado6",
    "Detalle: Personalizado7",
    "Detalle: Personalizado8",
    "Detalle: Personalizado9",
    "Detalle: Personalizado10",
    "Detalle: Personalizado11",
    "Detalle: Personalizado12",
    "Detalle: Personalizado13",
    "Detalle: Personalizado14",
    "Detalle: Personalizado15",
    "Detalle: Código Centro Costos",
]


def format_date_for_export(date_string: str) -> str:
    """Format date to DD/MM/YYYY format for World Office."""
    if not date_string:
        return ""
    # Extract YYYY-MM-DD from string (could have T suffix)
    date_part = date_string.split("T")[0]
    parts = date_part.split("-")
    if len(parts) == 3:
        year, month, day = parts
        return f"{day}/{month}/{year}"
    return date_string


def calculate_due_date(delivery_date_str: str, credit_days: int) -> str:
    """Calculate due date = delivery date + credit days."""
    if not delivery_date_str:
        return ""

    date_part = delivery_date_str.split("T")[0]
    parts = date_part.split("-")
    if len(parts) != 3:
        return ""

    year, month, day = map(int, parts)
    delivery_date = datetime(year, month, day)
    due_date = delivery_date + timedelta(days=credit_days)

    return f"{due_date.day:02d}/{due_date.month:02d}/{due_date.year}"


def get_world_office_config(supabase) -> Dict[str, Any]:
    """Get World Office configuration from system_config table."""
    config_keys = [
        "wo_company_name",
        "wo_document_type",
        "wo_document_prefix",
        "wo_third_party_internal",
        "wo_third_party_external",
        "wo_warehouse",
        "wo_unit_measure",
        "wo_iva_rate",
    ]

    result = (
        supabase.table("system_config")
        .select("config_key, config_value")
        .in_("config_key", config_keys)
        .execute()
    )

    config = {
        "company_name": "PAN",
        "document_type": "FV",
        "document_prefix": "PAN",
        "third_party_internal": "",
        "third_party_external": "",
        "warehouse": "BOD",
        "unit_measure": "UN",
        "iva_rate": 0.19,
    }

    for item in result.data:
        key = item["config_key"].replace("wo_", "")
        value = item["config_value"]
        if key == "iva_rate" and value:
            try:
                config[key] = float(value)
            except:
                pass
        else:
            config[key] = value or config.get(key, "")

    return config


def get_credit_terms(supabase, client_ids: List[str]) -> Dict[str, int]:
    """Get credit terms for clients."""
    if not client_ids:
        return {}

    result = (
        supabase.table("client_credit_terms")
        .select("client_id, credit_days")
        .in_("client_id", client_ids)
        .execute()
    )

    return {ct["client_id"]: ct["credit_days"] for ct in result.data}


def get_product_configs(supabase, product_ids: List[str]) -> Dict[str, Dict[str, Any]]:
    """Get product configurations (units_per_package)."""
    if not product_ids:
        return {}

    result = (
        supabase.table("product_configs")
        .select("product_id, units_per_package")
        .in_("product_id", product_ids)
        .execute()
    )

    return {pc["product_id"]: pc for pc in result.data}


def calculate_unit_price(
    package_price: float,
    units_per_package: int,
    product_id: str,
    client_id: str,
    client_price_lists: Dict[str, Dict[str, float]] = None
) -> float:
    """
    Calculate unit price.

    Priority:
    1. Client-specific price list
    2. Package price / units per package
    """
    # Check client-specific price list first
    if client_price_lists and client_id in client_price_lists:
        client_prices = client_price_lists[client_id]
        if product_id in client_prices:
            return client_prices[product_id]

    # Calculate from package price
    if units_per_package and units_per_package > 0:
        return package_price / units_per_package

    return package_price


def generate_world_office_excel(
    orders: List[Dict[str, Any]],
    items_by_order: Dict[str, List[Dict[str, Any]]],
    invoice_number_start: int,
    supabase,
) -> Dict[str, Any]:
    """
    Generate World Office Excel file.

    Args:
        orders: List of order dictionaries with client and branch info
        items_by_order: Dictionary mapping order_id to list of order items
        invoice_number_start: Starting invoice number
        supabase: Supabase client

    Returns:
        Dictionary with file_bytes and file_name
    """
    try:
        import openpyxl
        from openpyxl import Workbook
    except ImportError:
        logger.error("openpyxl not installed. Run: pip install openpyxl")
        raise ImportError("openpyxl library required for Excel generation")

    logger.info(f"Generating World Office Excel for {len(orders)} orders")

    # Get World Office config
    wo_config = get_world_office_config(supabase)

    # Get client IDs and product IDs
    client_ids = list(set(o.get("client_id") for o in orders if o.get("client_id")))
    all_product_ids = []
    for items in items_by_order.values():
        for item in items:
            if item.get("product_id"):
                all_product_ids.append(item["product_id"])
    product_ids = list(set(all_product_ids))

    # Get credit terms and product configs
    credit_terms = get_credit_terms(supabase, client_ids)
    product_configs = get_product_configs(supabase, product_ids)

    # Get client price lists
    client_price_lists = {}
    if client_ids:
        price_list_result = (
            supabase.table("client_price_lists")
            .select("client_id, product_id, unit_price")
            .in_("client_id", client_ids)
            .execute()
        )
        for pl in price_list_result.data:
            cid = pl["client_id"]
            if cid not in client_price_lists:
                client_price_lists[cid] = {}
            client_price_lists[cid][pl["product_id"]] = pl["unit_price"]

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Encab+Movim.Inven Talla y Color"

    # Write headers
    for col_idx, header in enumerate(EXCEL_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Generate rows
    row_num = 2
    current_invoice = invoice_number_start

    for order in orders:
        order_id = order["id"]
        items = items_by_order.get(order_id, [])

        if not items:
            continue

        client = order.get("clients") or {}
        branch = order.get("branches") or {}

        # Get credit days for this client
        client_id = order.get("client_id")
        credit_days = credit_terms.get(client_id, 30)  # Default 30 days

        # Format dates
        delivery_date = order.get("expected_delivery_date", "")
        delivery_date_formatted = format_date_for_export(delivery_date)
        due_date_formatted = calculate_due_date(delivery_date, credit_days)

        # Branch info for Encab: Sucursal
        branch_info = branch.get("name") or client.get("name") or ""

        # Get assigned user cedula for Tercero Interno
        tercero_interno = wo_config.get("third_party_internal", "")
        if client.get("assigned_user"):
            cedula = client["assigned_user"].get("cedula")
            if cedula:
                tercero_interno = cedula

        for item in items:
            product = item.get("products") or {}
            quantity_available = item.get("quantity_available") or item.get("quantity_requested") or 0

            if quantity_available <= 0:
                continue

            # Get units per package
            product_id = item.get("product_id")
            pc = product_configs.get(product_id, {})
            units_per_package = pc.get("units_per_package") or product.get("units_per_package") or 1

            # Convert quantity from packages to units
            quantity_in_units = quantity_available * units_per_package

            # Calculate unit price
            package_price = product.get("price") or item.get("unit_price") or 0
            unit_price = calculate_unit_price(
                package_price,
                units_per_package,
                product_id,
                client_id,
                client_price_lists
            )

            # Calculate IVA based on product tax_rate
            product_tax_rate = product.get("tax_rate") or 0
            iva_rate = 0 if product_tax_rate == 0 else product_tax_rate / 100

            # Product code for World Office
            product_code = product.get("codigo_wo") or product.get("name") or ""

            # Build row data
            row_data = [
                wo_config.get("company_name", ""),           # Empresa
                wo_config.get("document_type", ""),          # Tipo Documento
                wo_config.get("document_prefix", ""),        # Prefijo
                current_invoice,                              # Documento Número
                delivery_date_formatted,                     # Fecha
                tercero_interno,                             # Tercero Interno
                client.get("nit") or wo_config.get("third_party_external", ""),  # Tercero Externo
                "",                                          # Nota
                "Credito",                                   # FormaPago
                delivery_date_formatted,                     # Fecha Entrega
                None,                                        # Prefijo Documento Externo
                None,                                        # Número_Documento_Externo
                None,                                        # Verificado
                None,                                        # Anulado
                None,                                        # Personalizado 1
                order.get("purchase_order_number"),          # Personalizado 2
                None,                                        # Personalizado 3
                None,                                        # Personalizado 4
                None,                                        # Personalizado 5
                None,                                        # Personalizado 6
                None,                                        # Personalizado 7
                None,                                        # Personalizado 8
                None,                                        # Personalizado 9
                None,                                        # Personalizado 10
                None,                                        # Personalizado 11
                None,                                        # Personalizado 12
                None,                                        # Personalizado 13
                None,                                        # Personalizado 14
                None,                                        # Personalizado 15
                branch_info,                                 # Sucursal
                None,                                        # Clasificación
                product_code,                                # Producto
                wo_config.get("warehouse", ""),              # Bodega
                wo_config.get("unit_measure", ""),           # UnidadDeMedida
                quantity_in_units,                           # Cantidad
                iva_rate,                                    # IVA
                round(unit_price),                           # Valor Unitario
                0,                                           # Descuento
                due_date_formatted,                          # Vencimiento
                None,                                        # Nota
                None,                                        # Centro costos
                None,                                        # Personalizado1
                None,                                        # Personalizado2
                None,                                        # Personalizado3
                None,                                        # Personalizado4
                None,                                        # Personalizado5
                None,                                        # Personalizado6
                None,                                        # Personalizado7
                None,                                        # Personalizado8
                None,                                        # Personalizado9
                None,                                        # Personalizado10
                None,                                        # Personalizado11
                None,                                        # Personalizado12
                None,                                        # Personalizado13
                None,                                        # Personalizado14
                None,                                        # Personalizado15
                None,                                        # Código Centro Costos
            ]

            # Write row
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_num, column=col_idx, value=value)

            row_num += 1

        # Each order gets its own invoice number
        current_invoice += 1

    # Add secondary sheet with IVA
    iva_sheet = wb.create_sheet("Hoja1")
    iva_sheet.cell(row=1, column=1, value="IVA")
    iva_sheet.cell(row=2, column=1, value=wo_config.get("iva_rate", 0.19))

    # Generate filename
    today = datetime.now()
    file_name = f"WorldOffice_{today.strftime('%Y-%m-%d')}.xlsx"

    # Save to bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    file_bytes = buffer.read()

    logger.info(f"Generated Excel with {row_num - 2} rows")

    return {
        "file_bytes": file_bytes,
        "file_name": file_name,
    }
